import pandas as pd
import phonenumbers
from phonenumbers.phonenumberutil import NumberParseException
import os
from twilio.rest import Client


# ============ Twilio Configuration ============
ACCOUNT_SID = "US4389fef7fe2faaf9f4503d98158b30c8"
AUTH_TOKEN = "73aaafe134db1a4a270f5b80de96f3e4"

try:
    twilio_client = Client(ACCOUNT_SID, AUTH_TOKEN)
    TWILIO_ENABLED = True
except Exception as e:
    print(f"⚠️  Twilio not configured. Install: pip install twilio")
    TWILIO_ENABLED = False


def validate_phone_format(phone, country_code="GB"):
    """Validate phone number format using libphonenumber."""
    if pd.isna(phone) or not str(phone).strip():
        return False, None
    
    try:
        phone_str = str(phone).strip()
        parsed = phonenumbers.parse(phone_str, country_code)
        is_valid = phonenumbers.is_valid_number(parsed)
        
        if is_valid:
            formatted = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
            return True, formatted
        return False, None
        
    except (NumberParseException, Exception):
        return False, None


def verify_phone_exists(phone, country_code="GB"):
    """
    Verify if phone number actually exists using Twilio Lookup API.
    Returns: (exists, carrier_info, error_message)
    """
    if not TWILIO_ENABLED:
        return None, None, "Twilio not configured"
    
    try:
        # First validate format
        is_valid_format, formatted = validate_phone_format(phone, country_code)
        if not is_valid_format:
            return False, None, "Invalid format"
        
        # Use Twilio Lookup API
        lookup = twilio_client.lookups.v1.phone_numbers(formatted).fetch(type=['carrier'])
        
        carrier_name = lookup.carrier.get('name', 'Unknown')
        return True, carrier_name, None
        
    except Exception as e:
        error_msg = str(e)
        if "not found" in error_msg.lower():
            return False, None, "Number not found"
        elif "invalid" in error_msg.lower():
            return False, None, "Invalid number"
        else:
            return None, None, str(e)


def validate_file(input_file, country_code="GB", output_file=None, max_records=3000, use_twilio=True):
    """
    Validate phone numbers in Excel (.xlsx) or CSV file.
    
    Args:
        input_file: Path to input file (.xlsx or .csv)
        country_code: ISO country code (GB=UK, IN=India, US=USA, etc.)
        output_file: Path to save invalid phone numbers
        max_records: Max records to process (default 3000 for free tier)
        use_twilio: Use Twilio API to verify if number exists
    """
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"❌ Error: File '{input_file}' not found.")
        return
    
    # Determine file type
    file_ext = os.path.splitext(input_file)[1].lower()
    
    if file_ext not in ['.xlsx', '.csv']:
        print(f"❌ Error: Unsupported file type '{file_ext}'. Use .xlsx or .csv")
        return
    
    try:
        # Read file based on type
        print(f"📂 Reading {file_ext.upper()} file...")
        
        if file_ext == '.xlsx':
            df = pd.read_excel(input_file)
        else:  # .csv
            df = pd.read_csv(input_file, encoding='utf-8')
        
        # Check if 'phone' column exists
        if 'phone' not in df.columns:
            print(f"❌ Error: 'phone' column not found in file.")
            print(f"📋 Available columns: {', '.join(df.columns)}")
            return
        
        # Limit records to max_records
        if len(df) > max_records:
            print(f"⚠️  Limiting to {max_records} records for testing (free tier)")
            df = df.head(max_records)
        
        invalid_records = []
        valid_count = 0
        total_count = len(df)
        twilio_checked = 0
        
        country_names = {
            'GB': 'United Kingdom',
            'IN': 'India',
            'US': 'USA',
            'AU': 'Australia',
            'CA': 'Canada'
        }
        
        print(f"🌍 Country Code: {country_names.get(country_code, country_code)}")
        print(f"✓ Processing {total_count} records...")
        if use_twilio and TWILIO_ENABLED:
            print(f"🔍 Verifying with Twilio Lookup API...\n")
        else:
            print(f"⚠️  Twilio verification disabled\n")
        
        # Validate each row
        for row_num, (index, row) in enumerate(df.iterrows(), start=2):
            phone = row['phone']
            is_valid_format, formatted = validate_phone_format(phone, country_code)
            
            twilio_status = None
            carrier = None
            
            # If format is valid and Twilio enabled, verify existence
            if is_valid_format and use_twilio and TWILIO_ENABLED:
                exists, carrier, error = verify_phone_exists(phone, country_code)
                twilio_status = exists
                twilio_checked += 1
                
                # Print progress
                if twilio_checked % 100 == 0:
                    print(f"  ✓ Checked {twilio_checked} numbers with Twilio...")
            
            # Determine if number is invalid
            is_invalid = False
            reason = None
            
            if not is_valid_format:
                is_invalid = True
                reason = "Invalid format"
            elif use_twilio and TWILIO_ENABLED and twilio_status is False:
                is_invalid = True
                reason = "Number not found (Twilio)"
            elif use_twilio and TWILIO_ENABLED and twilio_status is None:
                is_invalid = True
                reason = "Verification error"
            
            # Store invalid records
            if is_invalid:
                invalid_records.append({
                    'row': f"r{row_num}",
                    'phone': phone,
                    'formatted': formatted or '',
                    'reason': reason,
                    'carrier': carrier or '',
                    'company_name': row.get('company_name', ''),
                    'address_1': row.get('address_1', ''),
                    'post_code': row.get('post_code', ''),
                    'post_town': row.get('post_town', '')
                })
            else:
                valid_count += 1
        
        # Generate output filename if not provided
        if output_file is None:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            output_file = f"{base_name}_invalid_phones.xlsx"
        
        # Save invalid records to Excel with red highlighting
        if invalid_records:
            invalid_df = pd.DataFrame(invalid_records)
            
            excel_file = output_file.replace('.csv', '.xlsx')
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                invalid_df.to_excel(writer, sheet_name='Invalid Phones', index=False)
                
                # Apply red color to invalid phone numbers
                from openpyxl.styles import PatternFill, Font
                workbook = writer.book
                worksheet = writer.sheets['Invalid Phones']
                
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                white_font = Font(bold=True, color='FFFFFF')
                
                # Color the phone column (column B, starting from row 2)
                for row in range(2, len(invalid_df) + 2):
                    cell = worksheet.cell(row=row, column=2)  # Column B is phone
                    cell.fill = red_fill
                    cell.font = white_font
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output_file = excel_file
        
        # Print summary report
        print(f"\n{'='*70}")
        print(f"{'PHONE VALIDATION REPORT':^70}")
        print(f"{'='*70}")
        print(f"File Type:              {file_ext.upper()}")
        print(f"Country:                {country_names.get(country_code, country_code)}")
        print(f"Records Processed:      {total_count}")
        print(f"Valid Numbers:          {valid_count} ✓")
        print(f"Invalid Numbers:        {len(invalid_records)} ✗")
        if use_twilio and TWILIO_ENABLED:
            print(f"Twilio Lookups Used:    {twilio_checked}")
        if total_count > 0:
            print(f"Valid Percentage:       {(valid_count/total_count*100):.2f}%")
        print(f"{'='*70}\n")
        
        if invalid_records:
            print(f"📊 Invalid phone numbers saved to: {output_file}\n")
            print(f"📋 First 15 Invalid Records:")
            print(f"{'-'*70}")
            for i, record in enumerate(invalid_records[:15], 1):
                print(f"  {i}. {record['row']}: {record['phone']:20} | {record['reason']}")
            print(f"{'-'*70}")
            if len(invalid_records) > 15:
                print(f"  ... and {len(invalid_records) - 15} more records")
        else:
            print(f"✅ All phone numbers are valid!")
        
    except Exception as e:
        print(f"❌ Error processing file: {str(e)}")


if __name__ == "__main__":
    print("\n🔍 PHONE NUMBER VALIDATOR WITH TWILIO\n")
    
    # Setup Twilio credentials
    print("⚙️  Setting up Twilio credentials...")
    print("Get your credentials from: https://www.twilio.com/console\n")
    
    ACCOUNT_SID = input("Enter your Twilio ACCOUNT_SID: ").strip()
    AUTH_TOKEN = input("Enter your Twilio AUTH_TOKEN: ").strip()
    
    if ACCOUNT_SID and AUTH_TOKEN:
        # Update global credentials
        globals()['ACCOUNT_SID'] = ACCOUNT_SID
        globals()['AUTH_TOKEN'] = AUTH_TOKEN
        globals()['twilio_client'] = Client(ACCOUNT_SID, AUTH_TOKEN)
        globals()['TWILIO_ENABLED'] = True
        print("✓ Twilio configured!\n")
    else:
        print("⚠️  Skipping Twilio verification (format validation only)\n")
    
    # Process file
    input_file = r'D:\VETRI-DQX-main\Company_Issues(Company_Issues) (1).xlsx'
    validate_file(input_file, country_code='GB', max_records=3000, use_twilio=True)